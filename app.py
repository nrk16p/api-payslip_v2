
import os
from dotenv import load_dotenv

load_dotenv()   # â† à¸•à¹‰à¸­à¸‡à¸­à¸¢à¸¹à¹ˆà¸•à¸£à¸‡à¸™à¸µà¹‰
import pandas as pd
from flask import Flask, jsonify, request , send_file
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from sqlalchemy import (
    create_engine, Column, Integer, String, DECIMAL, Enum, ForeignKey,
    TIMESTAMP, text, DateTime, Boolean
)
from sqlalchemy.orm import declarative_base, sessionmaker, scoped_session, relationship
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from functools import lru_cache
from werkzeug.utils import secure_filename
from flask_cors import CORS

# â”€â”€â”€ Timezone â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

TZ_BKK = ZoneInfo("Asia/Bangkok")

def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def as_utc(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

def utc_to_bkk(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_BKK)

# â”€â”€â”€ Configuration â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
DATABASE_URL = os.getenv("DATABASE_URL")
UPLOAD_DIR = "/tmp/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# â”€â”€â”€ Database Engine â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
engine = create_engine(
    DATABASE_URL,
    pool_size=5,
    max_overflow=5,
    pool_timeout=30,
    pool_recycle=300,
    pool_pre_ping=True,
    future=True,
)

SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
Session = scoped_session(SessionLocal)
Base = declarative_base()

# â”€â”€â”€ Models â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
class Employee(Base):
    __tablename__ = "employees"
    employee_id = Column(Integer, primary_key=True, autoincrement=True)
    emp_code = Column(String(64), unique=True, nullable=False)
    full_name = Column(String(255), nullable=False)
    status_name = Column(String(100), default="à¸›à¸à¸•à¸´")
    created_at = Column(TIMESTAMP, default=now_utc)
    items = relationship("SalaryItem", back_populates="employee")


class SalarySheet(Base):
    __tablename__ = "salary_sheets"
    sheet_id = Column(Integer, primary_key=True, autoincrement=True)
    month_year = Column(String(50), unique=True, nullable=False)

    api_active_from = Column(DateTime, nullable=True)  # UTC
    api_active_to = Column(DateTime, nullable=True)    # UTC
    api_is_active = Column(Boolean, default=False)

    created_at = Column(TIMESTAMP, default=now_utc)
    items = relationship("SalaryItem", back_populates="sheet")


class SalaryItem(Base):
    __tablename__ = "salary_items"
    item_id = Column(Integer, primary_key=True, autoincrement=True)
    sheet_id = Column(Integer, ForeignKey("salary_sheets.sheet_id"), nullable=False)
    employee_id = Column(Integer, ForeignKey("employees.employee_id"), nullable=False)
    item_group = Column(Enum("earnings", "deductions", "summary"), nullable=False)
    item_name = Column(String(255), nullable=False)
    amount = Column(DECIMAL(14, 2), default=0)

    sheet = relationship("SalarySheet", back_populates="items")
    employee = relationship("Employee", back_populates="items")


class SalaryItemMeta(Base):
    __tablename__ = "salary_item_meta"
    meta_id = Column(Integer, primary_key=True, autoincrement=True)
    item_name = Column(String(255), unique=True, nullable=False)
    item_group = Column(Enum("earnings", "deductions", "summary"))
    remark = Column(String(255))
    updated_at = Column(TIMESTAMP, default=now_utc, onupdate=now_utc)

class Salary50Tawi(Base):
    __tablename__ = "salary_50tawi"

    id = Column(Integer, primary_key=True, autoincrement=True)
    year = Column(String(10), nullable=False)  # 2569
    employee_id = Column(Integer, ForeignKey("employees.employee_id"), nullable=False)
    url_pdf = Column(String(500), nullable=True)

    created_at = Column(TIMESTAMP, default=now_utc)

    employee = relationship("Employee")

Base.metadata.create_all(bind=engine)

# â”€â”€â”€ Flask App â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
app = Flask(__name__)
CORS(
    app,
    resources={r"/*": {"origins": "*"}},
    supports_credentials=False,
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)

@app.teardown_appcontext
def remove_session(exception=None):
    Session.remove()

# â”€â”€â”€ Cache Helper â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@lru_cache(maxsize=256)
def load_item_meta():
    session = Session()
    rows = session.execute(
        text("SELECT item_name, item_group FROM salary_item_meta")
    ).fetchall()
    session.close()
    return {r[0]: r[1] for r in rows}

from datetime import timezone

def as_utc(dt: datetime) -> datetime:
    """Ensure datetime is UTC-aware"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

# â”€â”€â”€ Health Check â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/healthz")
def healthz():
    return jsonify({"status": "OK", "timezone": "Asia/Bangkok"}), 200

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 1ï¸âƒ£ GET & POST salary_data
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/salary_data/data", methods=["GET", "POST"])
def salary_data():
    session = Session()

    try:
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # GET : read + time guard
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if request.method == "GET":
            month = request.args.get("month-year")
            emp_code = request.args.get("emp_id")

            if not month or not emp_code:
                return jsonify({"error": "month-year and emp_id required"}), 400

            sheet = session.query(SalarySheet).filter_by(month_year=month).first()
            emp = session.query(Employee).filter_by(emp_code=emp_code).first()

            if not sheet or not emp:
                return jsonify([])

            now = now_utc()
            active_from = as_utc(sheet.api_active_from)
            active_to = as_utc(sheet.api_active_to)

            if active_from and now < active_from:
                return jsonify([]), 200


            if active_to and now > active_to:
                return jsonify([]), 200


            items = session.query(SalaryItem).filter_by(
                sheet_id=sheet.sheet_id,
                employee_id=emp.employee_id
            ).all()

            grouped = {"earnings": {}, "deductions": {}, "summary": {}}
            for i in items:
                grouped[i.item_group][i.item_name] = f"{float(i.amount):.2f}"

            return jsonify([{
                "Sheet": sheet.month_year,
                "à¸£à¸«à¸±à¸ªà¸à¸™à¸±à¸à¸‡à¸²à¸™": emp.emp_code,
                "à¸Šà¸·à¹ˆà¸­ - à¸™à¸²à¸¡à¸ªà¸à¸¸à¸¥": emp.full_name,
                "à¸ªà¸–à¸²à¸™à¸°à¸„à¸™à¸¥à¸²à¸­à¸­à¸": emp.status_name,
                "datalist": grouped,
            }])

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # POST : upsert salary (à¹„à¸¡à¹ˆà¸„à¸¸à¸¡à¹€à¸§à¸¥à¸²)
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        data = request.get_json(force=True)

        month = data.get("month-year")
        emp_code = data.get("emp_id")

        if not month or not emp_code:
            return jsonify({"error": "month-year and emp_id required"}), 400

        full_name = data.get("full_name", "")
        status = data.get("status", "à¸›à¸à¸•à¸´")
        datalist = data.get("datalist", {})

        sheet = session.query(SalarySheet).filter_by(month_year=month).first()
        if not sheet:
            sheet = SalarySheet(month_year=month)
            session.add(sheet)
            session.flush()

        session.execute(text("""
            INSERT INTO employees (emp_code, full_name, status_name, created_at)
            VALUES (:code, :name, :status, NOW())
            ON DUPLICATE KEY UPDATE full_name=:name, status_name=:status
        """), {"code": emp_code, "name": full_name, "status": status})

        emp = session.query(Employee).filter_by(emp_code=emp_code).first()

        session.query(SalaryItem).filter_by(
            sheet_id=sheet.sheet_id,
            employee_id=emp.employee_id
        ).delete()

        meta_map = load_item_meta()

        for group, items in datalist.items():
            for name, val in items.items():
                try:
                    amount = float(val)
                except:
                    continue

                session.add(SalaryItem(
                    sheet_id=sheet.sheet_id,
                    employee_id=emp.employee_id,
                    item_group=meta_map.get(name, group),
                    item_name=name,
                    amount=amount
                ))

        session.commit()
        return jsonify({"status": "updated"}), 201

    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        session.close()


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 2ï¸âƒ£ Upload Excel endpoint
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/upload_excel", methods=["POST"])
def upload_excel():

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ğŸ“‚ Validate file
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]

    if not file.filename:
        return jsonify({"error": "Empty filename"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_DIR, filename)
    file.save(filepath)

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ğŸ“– Read Excel
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        return jsonify({"error": f"Failed to read Excel: {e}"}), 400

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ğŸ§¹ Clean columns
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    df.columns = df.columns.astype(str).str.strip()
    df.columns = [c if not c.startswith("_") else f"Unnamed{c}" for c in df.columns]
    df = df.dropna(axis=1, how="all")

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ğŸ—“ Convert Thai month (à¸.à¸¢.2568 â†’ November2568)
    # (same as your V1 logic)
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    prefix_map = {
        "à¸¡.à¸„.": "January",
        "à¸.à¸.": "February",
        "à¸¡à¸µ.à¸„.": "March",
        "à¹€à¸¡.à¸¢.": "April",
        "à¸.à¸„.": "May",
        "à¸¡à¸´.à¸¢.": "June",
        "à¸.à¸„.": "July",
        "à¸ª.à¸„.": "August",
        "à¸.à¸¢.": "September",
        "à¸•.à¸„.": "October",
        "à¸.à¸¢.": "November",
        "à¸˜.à¸„.": "December"
    }

    if "Sheet" in df.columns:
        s = df["Sheet"].astype(str).str.replace(r"\s+", "", regex=True)
        df[["prefix", "year_th"]] = s.str.extract(r"^(\D+)(\d{4})$")
        df["Sheet"] = (
            df["prefix"].map(prefix_map).fillna(df["prefix"])
            + df["year_th"].astype(str)
        )

    month_value = str(df.iloc[0].get("Sheet", "Unknown")).strip()

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # ğŸ’¾ Start DB session
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    session = Session()
    inserted_rows = 0

    try:

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Ensure SalarySheet exists
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        sheet = session.query(SalarySheet).filter_by(
            month_year=month_value
        ).first()

        if not sheet:
            sheet = SalarySheet(month_year=month_value)
            session.add(sheet)
            session.flush()

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Load metadata
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        meta_rows = session.query(
            SalaryItemMeta.item_name,
            SalaryItemMeta.item_group
        ).all()

        meta_map = {row.item_name: row.item_group for row in meta_rows}

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # ğŸ”’ STRICT Salary Item Validation
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        TOP_LEVEL = [
            "Sheet",
            "à¸£à¸«à¸±à¸ªà¸à¸™à¸±à¸à¸‡à¸²à¸™",
            "à¸Šà¸·à¹ˆà¸­-à¸™à¸²à¸¡à¸ªà¸à¸¸à¸¥",
            "à¸ªà¸–à¸²à¸™à¸°à¸„à¸™à¸¥à¸²à¸­à¸­à¸",
            "prefix",
            "year_th"
        ]

        excel_salary_cols = [
            col for col in df.columns
            if col not in TOP_LEVEL
        ]

        unknown_cols = [
            col for col in excel_salary_cols
            if col not in meta_map
        ]

        if unknown_cols:
            session.close()
            return jsonify({
                "error": "Unknown salary items detected",
                "message": "Some Excel columns do not match salary_item_meta.",
                "unknown_columns": unknown_cols,
                "allowed_columns": sorted(list(meta_map.keys())),
                "hint": "Please fix spelling or create metadata before uploading."
            }), 400

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Preload employees
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        emp_rows = session.query(
            Employee.emp_code,
            Employee.employee_id
        ).all()

        emp_map = {e.emp_code: e.employee_id for e in emp_rows}

        salary_items = []
        batch_size = 10  # SAME AS V1

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # ğŸ” Iterate employees (UNCHANGED V1 LOGIC)
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        for _, row in df.iterrows():

            emp_code = str(row.get("à¸£à¸«à¸±à¸ªà¸à¸™à¸±à¸à¸‡à¸²à¸™", "")).strip()
            full_name = str(row.get("à¸Šà¸·à¹ˆà¸­-à¸™à¸²à¸¡à¸ªà¸à¸¸à¸¥", "")).strip()
            status = str(row.get("à¸ªà¸–à¸²à¸™à¸°à¸„à¸™à¸¥à¸²à¸­à¸­à¸", "à¸›à¸à¸•à¸´")).strip()

            if not emp_code or emp_code.lower() in ["nan", "none"]:
                continue

            # Upsert employee
            emp_id = emp_map.get(emp_code)

            if not emp_id:
                session.execute(text("""
                    INSERT INTO employees (emp_code, full_name, status_name, created_at)
                    VALUES (:code, :name, :status, NOW())
                    ON DUPLICATE KEY UPDATE
                        full_name = :name,
                        status_name = :status
                """), {
                    "code": emp_code,
                    "name": full_name,
                    "status": status
                })

                session.flush()

                emp = session.query(Employee).filter_by(
                    emp_code=emp_code
                ).first()

                emp_id = emp.employee_id
                emp_map[emp_code] = emp_id

            # Delete existing salary items for this employee + sheet
            session.query(SalaryItem).filter_by(
                sheet_id=sheet.sheet_id,
                employee_id=emp_id
            ).delete()

            # Build salary items
            for col in excel_salary_cols:

                val = row.get(col)

                if pd.isna(val):
                    continue

                try:
                    amount = float(val)
                except Exception:
                    continue

                salary_items.append({
                    "sheet_id": sheet.sheet_id,
                    "employee_id": emp_id,
                    "item_group": meta_map[col],
                    "item_name": col,
                    "amount": amount,
                })

                inserted_rows += 1

            # Batch commit (same as V1)
            if inserted_rows % batch_size == 0:
                session.bulk_insert_mappings(
                    SalaryItem,
                    salary_items
                )
                salary_items.clear()
                session.commit()

        # Commit remaining
        if salary_items:
            session.bulk_insert_mappings(
                SalaryItem,
                salary_items
            )
            session.commit()

    except Exception as e:
        session.rollback()
        return jsonify({
            "error": f"DB error: {str(e)}"
        }), 500

    finally:
        session.close()
        load_item_meta.cache_clear()

    return jsonify({
        "status": "success",
        "sheet": month_value,
        "rows_inserted": inserted_rows
    }), 201

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 3ï¸âƒ£ salary_items/meta CRUD
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/salary_items/meta", methods=["GET", "POST", "DELETE"])
def salary_item_meta():
    session = Session()

    if request.method == "GET":
        rows = session.execute(
            text("SELECT meta_id, item_name, item_group, remark, updated_at FROM salary_item_meta ORDER BY item_name ASC")
        ).fetchall()
        session.close()
        return jsonify([
            {
                "meta_id": r[0],
                "item_name": r[1],
                "item_group": r[2],
                "remark": r[3],
                "updated_at": r[4].strftime("%Y-%m-%d %H:%M:%S"),
            } for r in rows
        ])

    if request.method == "POST":
        data = request.get_json(force=True)
        name = data.get("item_name")
        group = data.get("item_group")
        remark = data.get("remark", "")
        if not name or group not in ["earnings", "deductions", "summary"]:
            session.close()
            return jsonify({"error": "Invalid payload"}), 400

        session.execute(
            text("""
                INSERT INTO salary_item_meta (item_name, item_group, remark)
                VALUES (:name, :group, :remark)
                ON DUPLICATE KEY UPDATE item_group=:group, remark=:remark
            """),
            {"name": name, "group": group, "remark": remark},
        )
        session.commit()
        session.close()
        load_item_meta.cache_clear()
        return jsonify({"status": "updated", "item_name": name, "item_group": group}), 201

    if request.method == "DELETE":
        data = request.get_json(force=True)
        name = data.get("item_name")
        if not name:
            session.close()
            return jsonify({"error": "item_name required"}), 400

        result = session.execute(
            text("DELETE FROM salary_item_meta WHERE item_name=:name"), {"name": name}
        )
        session.commit()
        session.close()
        load_item_meta.cache_clear()

        if result.rowcount == 0:
            return jsonify({"status": "not_found", "item_name": name}), 404
        return jsonify({"status": "deleted", "item_name": name}), 200


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# API: Set API Window (Bangkok time input)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/salary_sheets/api-window", methods=["PATCH", "OPTIONS"])
def set_api_window():
    # âœ… Handle CORS preflight
    if request.method == "OPTIONS":
        return jsonify({}), 200

    session = Session()

    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"error": "invalid json body"}), 400

        sheet_id = data.get("sheet_id")
        if not sheet_id:
            return jsonify({"error": "sheet_id required"}), 400

        sheet = session.get(SalarySheet, sheet_id)
        if not sheet:
            return jsonify({"error": "sheet not found"}), 404

        # --- api_is_active ---
        if "api_is_active" in data:
            sheet.api_is_active = bool(data["api_is_active"])

        # --- api_active_from ---
        if "api_active_from" in data:
            try:
                sheet.api_active_from = (
                    datetime.fromisoformat(data["api_active_from"])
                    .replace(tzinfo=TZ_BKK)
                    .astimezone(timezone.utc)
                )
            except Exception as e:
                return jsonify({
                    "error": "invalid api_active_from",
                    "detail": str(e)
                }), 400

        # --- api_active_to ---
        if "api_active_to" in data:
            try:
                sheet.api_active_to = (
                    datetime.fromisoformat(data["api_active_to"])
                    .replace(tzinfo=TZ_BKK)
                    .astimezone(timezone.utc)
                )
            except Exception as e:
                return jsonify({
                    "error": "invalid api_active_to",
                    "detail": str(e)
                }), 400

        session.commit()

        # âœ… build response BEFORE close session
        resp = {
            "sheet_id": sheet.sheet_id,
            "api_is_active": sheet.api_is_active,
            "api_active_from_bkk": (
                utc_to_bkk(sheet.api_active_from).isoformat()
                if sheet.api_active_from else None
            ),
            "api_active_to_bkk": (
                utc_to_bkk(sheet.api_active_to).isoformat()
                if sheet.api_active_to else None
            ),
        }

        return jsonify(resp), 200

    except Exception as e:
        session.rollback()
        return jsonify({
            "error": "internal server error",
            "detail": str(e)
        }), 500

    finally:
        session.close()
@app.route("/salary_sheets/api-window", methods=["GET"])
def get_api_window():
    session = Session()
    try:
        sheet_id = request.args.get("sheet_id")
        month = request.args.get("month-year")

        q = session.query(SalarySheet)

        if sheet_id:
            q = q.filter(SalarySheet.sheet_id == sheet_id)
        elif month:
            q = q.filter(SalarySheet.month_year == month)

        sheets = q.order_by(SalarySheet.created_at.desc()).all()

        result = []
        now = now_utc()

        for s in sheets:
            active_from = as_utc(s.api_active_from)
            active_to = as_utc(s.api_active_to)

            is_active_now = (
                s.api_is_active
                and (not active_from or now >= active_from)
                and (not active_to or now <= active_to)
            )

            result.append({
                "sheet_id": s.sheet_id,
                "month_year": s.month_year,
                "api_is_active": s.api_is_active,
                "api_active_from_bkk": utc_to_bkk(active_from) if active_from else None,
                "api_active_to_bkk": utc_to_bkk(active_to) if active_to else None,
                "is_active_now": is_active_now
            })

        return jsonify(result), 200

    finally:
        session.close()

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# 4ï¸âƒ£ 50à¸—à¸§à¸´ API
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/50tawi/data", methods=["GET", "POST"])
def salary_50tawi():
    session = Session()

    try:
        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ GET â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        if request.method == "GET":
            year = request.args.get("year")
            emp_code = request.args.get("emp_id")

            if not year or not emp_code:
                return jsonify({"error": "year and emp_id required"}), 400

            emp = session.query(Employee).filter_by(emp_code=emp_code).first()
            if not emp:
                return jsonify([])

            record = session.query(Salary50Tawi).filter_by(
                year=year,
                employee_id=emp.employee_id
            ).first()

            if not record:
                return jsonify([])

            return jsonify([{
                "Sheet": year,
                "url_pdf": record.url_pdf,
                "à¸Šà¸·à¹ˆà¸­ - à¸™à¸²à¸¡à¸ªà¸à¸¸à¸¥": emp.full_name,
                "à¸£à¸«à¸±à¸ªà¸à¸™à¸±à¸à¸‡à¸²à¸™": emp.emp_code,
                "à¸ªà¸–à¸²à¸™à¸°à¸„à¸™à¸¥à¸²à¸­à¸­à¸": emp.status_name,
            }])

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ POST (upsert) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        data = request.get_json(force=True)

        year = data.get("year")
        emp_code = data.get("emp_id")
        url_pdf = data.get("url_pdf")

        if not year or not emp_code:
            return jsonify({"error": "year and emp_id required"}), 400

        emp = session.query(Employee).filter_by(emp_code=emp_code).first()
        if not emp:
            return jsonify({"error": "employee not found"}), 404

        record = session.query(Salary50Tawi).filter_by(
            year=year,
            employee_id=emp.employee_id
        ).first()

        if not record:
            record = Salary50Tawi(
                year=year,
                employee_id=emp.employee_id,
                url_pdf=url_pdf
            )
            session.add(record)
        else:
            record.url_pdf = url_pdf

        session.commit()

        return jsonify({"status": "updated"}), 201

    except Exception as e:
        session.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        session.close()

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Export Salary to Excel
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/salary_data/export", methods=["GET"])
def export_salary_month_pivot():
    session = Session()

    try:
        month = request.args.get("month-year")
        if not month:
            return jsonify({"error": "month-year required"}), 400

        sheet = session.query(SalarySheet).filter_by(month_year=month).first()
        if not sheet:
            return jsonify({"error": "sheet not found"}), 404

        rows = (
            session.query(
                Employee.emp_code,
                Employee.full_name,
                Employee.status_name,
                SalaryItem.item_group,
                SalaryItem.item_name,
                SalaryItem.amount,
            )
            .join(SalaryItem, SalaryItem.employee_id == Employee.employee_id)
            .filter(SalaryItem.sheet_id == sheet.sheet_id)
            .all()
        )

        if not rows:
            return jsonify({"error": "no salary data"}), 404

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€ Build DataFrame â”€â”€â”€â”€â”€â”€â”€â”€â”€
        data = []
        for r in rows:
            data.append({
                "à¸£à¸«à¸±à¸ªà¸à¸™à¸±à¸à¸‡à¸²à¸™": r.emp_code,
                "à¸Šà¸·à¹ˆà¸­ - à¸™à¸²à¸¡à¸ªà¸à¸¸à¸¥": r.full_name,
                "à¸ªà¸–à¸²à¸™à¸°": r.status_name,
                "à¸«à¸¡à¸§à¸”à¸«à¸¡à¸¹à¹ˆ": r.item_group,
                "à¸£à¸²à¸¢à¸à¸²à¸£": r.item_name,
                "à¸ˆà¸³à¸™à¸§à¸™à¹€à¸‡à¸´à¸™": float(r.amount),
            })

        df = pd.DataFrame(data)

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€ Pivot â”€â”€â”€â”€â”€â”€â”€â”€â”€
        pivot = df.pivot_table(
            index=["à¸£à¸«à¸±à¸ªà¸à¸™à¸±à¸à¸‡à¸²à¸™", "à¸Šà¸·à¹ˆà¸­ - à¸™à¸²à¸¡à¸ªà¸à¸¸à¸¥", "à¸ªà¸–à¸²à¸™à¸°"],
            columns="à¸£à¸²à¸¢à¸à¸²à¸£",
            values="à¸ˆà¸³à¸™à¸§à¸™à¹€à¸‡à¸´à¸™",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€ Calculate Net Pay â”€â”€â”€â”€â”€â”€â”€â”€â”€
        earnings_cols = df[df["à¸«à¸¡à¸§à¸”à¸«à¸¡à¸¹à¹ˆ"] == "earnings"]["à¸£à¸²à¸¢à¸à¸²à¸£"].unique()
        deductions_cols = df[df["à¸«à¸¡à¸§à¸”à¸«à¸¡à¸¹à¹ˆ"] == "deductions"]["à¸£à¸²à¸¢à¸à¸²à¸£"].unique()

        pivot["Total Earnings"] = pivot[list(earnings_cols)].sum(axis=1) if len(earnings_cols) else 0
        pivot["Total Deductions"] = pivot[list(deductions_cols)].sum(axis=1) if len(deductions_cols) else 0
        pivot["Net Pay"] = pivot["Total Earnings"] - pivot["Total Deductions"]

        # â”€â”€â”€â”€â”€â”€â”€â”€â”€ Write Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pivot.to_excel(writer, index=False, sheet_name="Payroll")

            worksheet = writer.sheets["Payroll"]

            # Bold header
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            # Auto column width
            for i, col in enumerate(pivot.columns, 1):
                max_length = max(
                    pivot[col].astype(str).map(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(i)].width = max_length + 5

        output.seek(0)

        filename = f"payroll_{month}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        session.close()

@app.route("/salary/employees", methods=["GET"])
def get_unique_employees():
    session = Session()
    try:
        results = (
            session.query(
                Employee.emp_code,
                Employee.full_name
            )
            .filter(Employee.full_name.isnot(None))   # âœ… à¸à¸±à¸™ NULL
            .filter(Employee.full_name != "")         # âœ… à¸à¸±à¸™à¸„à¹ˆà¸²à¸§à¹ˆà¸²à¸‡
            .distinct()
            .order_by(Employee.emp_code)
            .all()
        )

        data = [
            {
                "emp_code": r.emp_code,
                "full_name": r.full_name
            }
            for r in results
        ]

        return jsonify({
            "employees": data
        })

    finally:
        session.close()
@app.route("/salary/month-years", methods=["GET"])
def get_unique_month_years():
    session = Session()
    try:
        results = (
            session.query(SalarySheet.month_year)
            .distinct()
            .order_by(SalarySheet.month_year.desc())
            .all()
        )

        data = [r[0] for r in results]

        return jsonify({
            "month_years": data
        })

    finally:
        session.close()

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")
